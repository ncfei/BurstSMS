# The Program FP_ReplySMS.py obtain the message id then, get the SMS response list from BurstSMS. Then, it updates
# the master excel file with confirmed if the response from the SMS is not a no (meaning response is Yes or others).

# Program Written by: Chooi Fei Ng
# Email: fei.ngc@gmail.com
import requests
import BurstSMS_config
import openpyxl

class ExtractSMS:

    def __init__(self):
        # Initialise class
        self.message_id = 0
        self.response = []

    def mess_id(self):
        # Open and read the file message_id.txt to read the SMS message id
        f = open('message_id.txt', 'r')
        self.message_id = int(f.readlines()[0])
        #print(self.message_id)

    def get_response(self,pagenum):
        # Get sms response for messages with message_id from burst SMS website using POST request
        url = "https://api.transmitsms.com/get-sms-responses.json"
        files = []
        # Authorisation for BurstSMS account
        headers = {
            'Authorization': BurstSMS_config.apiKey
        }
        payload = {'message_id': self.message_id, 'page': pagenum}
        # Post request to url
        response = requests.request("POST", url, headers=headers, data=payload, files=files)
        if (response.status_code != 200):
            # If the status code is not 200, print out error messages
            print("Error FP_ReplySMS 01: Unable to retrieve Reply to SMS messages")
        # Save the json response in self.response
        res_json = response.json()
        return res_json
    def response_multipage(self):
        # This function go through each page of the POST request
        # Starting page number
        pagenum = 1
        pagecount = 1
        while pagenum <= pagecount:
            # Go through each page and get the responses for each page.
            resp = self.get_response(pagenum)
            #print(len(resp))
            #print(resp)
            # The total pages
            pagecount = resp["page"]["count"]
            # Current page number
            pagenum = pagenum + 1
            #print("pagenumber=",pagenum)
            #print("Pagecount =",pagecount)
            self.build_response_list(resp)
        # Return the self.response list
        return self.response
    def build_response_list(self,res):
        # This function build the response list and remove all the duplicates
        # Number of responses for each POST request
        res_count = len(res["responses"])
        #print("Number of responses: ", res_count)

        for i in range(res_count):
        # Go through the responses
            if(len(self.response) == 0):
                # First response ever, add into list of self.response
                self.response.append({"firstname":res["responses"][i]["first_name"],"lastname":res["responses"][i]["last_name"],"response":res["responses"][i]["response"]})
            else:
                # Check whether the firstname and lastname is already available in the self.response list
                firstname = res["responses"][i]["first_name"]
                lastname = res["responses"][i]["last_name"]
                found = False
                for j in range(len(self.response)):
                    # Go through the self.response list and found that the firstname and lastname already exist
                    if(self.response[j]["firstname"] == firstname and self.response[j]["lastname"] == lastname):
                        found = True

                if(found == False):
                    # If the name is not found, add it to the self.response list
                    self.response.append({"firstname": res["responses"][i]["first_name"],"lastname": res["responses"][i]["last_name"],"response": res["responses"][i]["response"]})
        #print(self.response)
        #print(len(self.response))
    def __str__(self):
        print(f"The SMS reply list is :\n", self.response)
        return "The SMS reply list count is: " + str(len(self.response))

class Update_Excel():
    def __init__(self,response):
        # Initialise class
        self.response = response
        #print(self.response)

    def change_excel(self):
        # This function search through the excel workbook and then updates the matching cells that doesnt have a no response
        # Open the excel file
        wb_obj = openpyxl.load_workbook("Food-Parcel Master List.xlsx")
        # Opens the sheet that is active
        sheet_obj = wb_obj.active
        #print(sheet_obj)
        # Determine that max column and row that has data in the excel file
        max_column = sheet_obj.max_column
        max_row = sheet_obj.max_row
        #print(max_column)
        #print(max_row)
        for i in range(len(self.response)):
            # Go through each of the response list
            for j in range(2,max_row - 3):
                # Go through each row of the excel file
                #print(i)
                # Read the first name and last name of each row
                firstname_cell = sheet_obj.cell(j,2)
                #print(firstname_cell.value)
                lastname_cell = sheet_obj.cell(j,3)
                #print(lastname_cell.value)
                if (self.response[i]["firstname"] == firstname_cell.value and self.response[i]["lastname"] == lastname_cell.value):
                    # Found the corresponding person in the self.response that is the same as the one in the excel file
                    if(self.response[i]["response"].lower() == 'no' or self.response[i]["response"].lower() == 'n' ):
                        # Do nothing, if the response is no
                        dummy = False
                        #print(firstname_cell.value)
                        #print(lastname_cell.value)
                        #print(response_cell.value)
                    else:
                        # If the response is yes or others, update the excel file to confirmed
                        response_cell = sheet_obj.cell(j,max_column)
                        response_cell.value = "Confirmed"
                        #print(firstname_cell.value)
                        #print(lastname_cell.value)
                        #print(response_cell.value)
        # Save the excel file
        try:
            wb_obj.save("Food-Parcel Master List.xlsx")
        except:
            print('Error FP_ReplySMS 02: Unable write to master file')

# Call the class and functions
# Initialise the class ExtractSMS
RetrieveSMS = ExtractSMS()
# Retrieve message id
RetrieveSMS.mess_id()
# Obtain the list of SMS response from burstSMS
SMS_res = RetrieveSMS.response_multipage()
#print(RetrieveSMS)
# Initialise the class Update Excel with the list of SMS response
Up_Excel = Update_Excel(SMS_res)
# Update the excel file with the confirmed response
Up_Excel.change_excel()

